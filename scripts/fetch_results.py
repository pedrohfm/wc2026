"""
Auto-fetch finished World Cup results from football-data.org (v4) and write them
into wc2026_results.xlsx — ONLY into blank score cells, never overwriting what you
typed.

Why this is safe and correct
----------------------------
* The engine keys results by MATCH NUMBER, so the job is: map each finished API
  match to our match number, then write Home/Away goals (+ a PK winner for
  shootouts) in OUR home/away orientation.
* Group fixtures map by team pair (order-independent). Knockout fixtures are
  resolved from results already entered, using the official third-place
  allocation (wc2026.config.THIRD_OVERRIDE) — so as earlier rounds fill in, later
  rounds become mappable. The script iterates until nothing new resolves, so one
  run can catch up several rounds.
* Scores: per the football-data docs, fullTime INCLUDES shootout goals, so for a
  PENALTY_SHOOTOUT we write the pre-shootout level score and set PK Win from the
  winner. EXTRA_TIME/REGULAR write fullTime directly.

Setup (token is a secret — never commit it):
    export FOOTBALL_DATA_TOKEN=your_token_here
Run:
    python scripts/fetch_results.py            # fetch + fill blanks + save
    python scripts/fetch_results.py --dry-run  # show what it would write
    python scripts/fetch_results.py --overwrite  # also CORRECT cells that disagree
                                                 # with the API (fixes bad entries)
    python scripts/fetch_results.py --selftest # offline parser checks, no network

--overwrite is the reconcile mode: where a score already in the file differs from
the authoritative API result, it replaces it (and reports the change). Use it to
fix a wrong entry; the default (blank-only) protects manual edits.
"""
from __future__ import annotations
import os, sys, json, unicodedata, urllib.request, urllib.error

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
sys.path.insert(0, os.path.join(ROOT, "src"))
import wc2026 as E
from wc2026.structure import GROUP_FIXTURES, KO
from wc2026.elo_dynamics import _deterministic_groups

XLSX = os.path.join(ROOT, "wc2026_results.xlsx")
ELO_CSV = os.path.join(ROOT, "wc2026_elo.csv")
KO_PENS = os.path.join(ROOT, "outputs", "ko_penalties.json")
BASE = "https://api.football-data.org/v4"
COMP = "WC"

# football-data team name -> our team name (extends the odds-fetcher's map)
NAME_MAP = {
    "Turkey": "Türkiye", "Turkiye": "Türkiye",
    "Ivory Coast": "Côte d'Ivoire", "Cote d'Ivoire": "Côte d'Ivoire",
    "Czech Republic": "Czechia",
    "Cape Verde": "Cabo Verde", "Cabo Verde Islands": "Cabo Verde", "Cape Verde Islands": "Cabo Verde",
    "Curacao": "Curaçao",
    "Bosnia and Herzegovina": "Bosnia & Herzegovina", "Bosnia-Herzegovina": "Bosnia & Herzegovina",
    "Congo DR": "DR Congo", "Democratic Republic of the Congo": "DR Congo", "DR Congo": "DR Congo",
    "Korea Republic": "South Korea", "Republic of Korea": "South Korea",
    "USA": "United States", "United States of America": "United States",
    "IR Iran": "Iran",
}
OURS = {t for (h, a, _g) in GROUP_FIXTURES.values() for t in (h, a)}


def _canon(s):
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode().lower()
    return "".join(ch for ch in s if ch.isalnum())


_CANON_OURS = {_canon(t): t for t in OURS}
_CANON_MAP = {_canon(k): v for k, v in NAME_MAP.items()}


def to_ours(name):
    """Map a football-data team name to our canonical team name, or None."""
    if name in NAME_MAP:
        return NAME_MAP[name]
    if name in OURS:
        return name
    c = _canon(name)
    return _CANON_OURS.get(c) or _CANON_MAP.get(c)


def _ha(node):
    """Read (home, away) from a score sub-node, tolerating both key styles."""
    if not isinstance(node, dict):
        return (None, None)
    return (node.get("home", node.get("homeTeam")),
            node.get("away", node.get("awayTeam")))


def parse_score(score, direct):
    """Return (home_goals, away_goals, pk) in OUR orientation, or None.
       `direct` is True when our home team is the API's home team.
       pk is 'H'/'A' (our orientation) only for shootouts, else None."""
    if not isinstance(score, dict):
        return None
    dur = score.get("duration") or "REGULAR"
    fth, fta = _ha(score.get("fullTime"))
    if fth is None or fta is None:
        return None
    if dur == "PENALTY_SHOOTOUT":
        ph, pa = _ha(score.get("penalties"))
        if ph is not None and pa is not None:
            gh, ga = fth - ph, fta - pa            # fullTime includes the shootout
        else:                                       # fallback: regular + extra time
            rh, ra = _ha(score.get("regularTime")); eh, ea = _ha(score.get("extraTime"))
            gh, ga = (rh or 0) + (eh or 0), (ra or 0) + (ea or 0)
        win = score.get("winner")
        pk = "H" if win == "HOME_TEAM" else ("A" if win == "AWAY_TEAM" else None)
    else:
        gh, ga, pk = fth, fta, None
    try:
        gh, ga = int(gh), int(ga)
    except (TypeError, ValueError):
        return None
    if direct:
        return (gh, ga, pk)
    return (ga, gh, ("A" if pk == "H" else "H" if pk == "A" else None))


def parse_pens(score, direct):
    """Penalty-shootout score as 'home-away' in OUR orientation, or None."""
    if not isinstance(score, dict) or score.get("duration") != "PENALTY_SHOOTOUT":
        return None
    ph, pa = _ha(score.get("penalties"))
    if ph is None or pa is None:
        return None
    return f"{int(ph)}-{int(pa)}" if direct else f"{int(pa)}-{int(ph)}"


def resolved_fixtures():
    """{(home, away): match_no} for every match currently determined by results,
       in OUR structure's home/away orientation. Plus (kg, kk)."""
    kg, kk = E.load_results(XLSX)
    fx = {(h, a): m for m, (h, a, _g) in GROUP_FIXTURES.items()}
    elo = E.load_elo(ELO_CSV)
    e = E.apply_known_results(elo, kg, kk, E.THIRD_OVERRIDE)
    res = _deterministic_groups(e, kg)
    if res:
        winners, runners, thirds = res
        thirds.sort(key=lambda x: (x[2]["pts"], x[2]["gd"], x[2]["gf"], e[x[1]]), reverse=True)
        qual = {g: t for g, t, _ in thirds[:8]}
        slot_group = dict(E.THIRD_OVERRIDE) if E.THIRD_OVERRIDE else E.allocate_thirds(list(qual))
        slot_team = {m: qual.get(g) for m, g in slot_group.items()}
        mwin, mlose = {}, {}
        for m in sorted(KO):
            rnd, hs, as_ = KO[m]
            def part(slot):
                typ, ref = slot
                return {"W": winners.get(ref), "RU": runners.get(ref), "3": slot_team.get(ref),
                        "WIN": mwin.get(ref), "LOSE": mlose.get(ref)}[typ]
            home, away = part(hs), part(as_)
            if home and away:
                fx[(home, away)] = m
                if m in kk:
                    ga, gb, pk = kk[m]
                    if ga > gb: w = home
                    elif gb > ga: w = away
                    else: w = home if pk == "H" else (away if pk == "A" else (home if e.get(home, 0) >= e.get(away, 0) else away))
                    mwin[m] = w; mlose[m] = away if w == home else home
    return fx, kg, kk


def api_get(path):
    token = os.environ.get("FOOTBALL_DATA_TOKEN", "").strip()
    if not token:
        raise RuntimeError("FOOTBALL_DATA_TOKEN is not set. Run: export FOOTBALL_DATA_TOKEN=your_token")
    req = urllib.request.Request(BASE + path, headers={"X-Auth-Token": token})
    try:
        with urllib.request.urlopen(req, timeout=30) as r:
            data = json.load(r)
            return data, dict(r.headers)
    except urllib.error.HTTPError as ex:
        if ex.code == 429:
            raise RuntimeError("Rate limited (free tier = 10 requests/min). Wait a minute and retry.")
        if ex.code in (401, 403):
            raise RuntimeError(f"Auth/permission error ({ex.code}) — check FOOTBALL_DATA_TOKEN and that WC is in your plan.")
        raise RuntimeError(f"HTTP {ex.code} from football-data: {ex.reason}")


def fetch_finished():
    season = os.environ.get("FOOTBALL_DATA_SEASON", "").strip()
    path = f"/competitions/{COMP}/matches?status=FINISHED" + (f"&season={season}" if season else "")
    data, headers = api_get(path)
    rem = headers.get("X-Requests-Available") or headers.get("X-RequestsAvailable")
    return data.get("matches", []), rem


def _open_sheet():
    import openpyxl
    wb = openpyxl.load_workbook(XLSX)          # default = keep formulas (Status col)
    ws = wb["Results"]
    hdr = next((r for r in range(1, 12)
                if any(str(ws.cell(r, c).value or "").strip().lower().startswith("match")
                       for c in range(1, ws.max_column + 1))), None)
    if hdr is None:
        raise RuntimeError("Couldn't find the header row in the Results sheet.")
    def find(*names):
        for c in range(1, ws.max_column + 1):
            v = str(ws.cell(hdr, c).value or "").strip().lower()
            if any(v == n or v.startswith(n) for n in names):
                return c
        return None
    cols = {"match": find("match"), "hg": find("home goals", "home_goals"),
            "ag": find("away goals", "away_goals"), "pk": find("pk")}
    if not all(cols[k] for k in ("match", "hg", "ag")):
        raise RuntimeError("Couldn't locate Match #/Home Goals/Away Goals columns.")
    row_of = {}
    for r in range(hdr + 1, ws.max_row + 1):
        try:
            row_of[int(float(ws.cell(r, cols["match"]).value))] = r
        except (TypeError, ValueError):
            continue
    return wb, ws, cols, row_of


def run(dry=False, overwrite=False):
    matches, rem = fetch_finished()
    print(f"  football-data: {len(matches)} finished WC match(es) returned"
          + (f"  ({rem} API requests left this minute)" if rem else ""))
    if not matches:
        print("  Nothing finished yet — nothing to write.")
        return
    wb, ws, cols, row_of = _open_sheet()
    filled, unmatched, total, pens_map = 0, set(), 0, {}
    for _pass in range(1 if dry else 8):         # iterate so later rounds resolve
        fx, kg, kk = resolved_fixtures()
        wrote = 0
        for am in matches:
            hn = (am.get("homeTeam") or {}).get("name")
            an = (am.get("awayTeam") or {}).get("name")
            oh, oa = to_ours(hn or ""), to_ours(an or "")
            if not oh or not oa:
                unmatched.add(f"{hn} vs {an}"); continue
            if (oh, oa) in fx:
                m, direct = fx[(oh, oa)], True
            elif (oa, oh) in fx:
                m, direct = fx[(oa, oh)], False
            else:
                continue                          # teams not yet resolvable
            row = row_of.get(m)
            if row is None:
                continue
            if m >= 73:                           # capture shootout score for the card
                pp = parse_pens(am.get("score") or {}, direct)
                if pp:
                    pens_map[str(m)] = pp
            parsed = parse_score(am.get("score") or {}, direct)
            if parsed is None:
                continue
            gh, ga, pk = parsed
            uh, ua = (oh, oa) if direct else (oa, oh)     # our home/away orientation
            cur_h, cur_a = ws.cell(row, cols["hg"]).value, ws.cell(row, cols["ag"]).value
            filled = cur_h not in (None, "") or cur_a not in (None, "")
            if filled:
                try:
                    same = int(float(cur_h)) == gh and int(float(cur_a)) == ga
                except (TypeError, ValueError):
                    same = False
                if same or not overwrite:
                    continue                      # matches, or protect manual entry
                tag = f"~ M{m}: {uh} {cur_h}-{cur_a} {ua}  ->  {gh}-{ga}"
            else:
                tag = f"+ M{m}: {uh} {gh}-{ga} {ua}"
            print(f"    {tag}" + (f" (PK {pk})" if pk else ""))
            if not dry:
                ws.cell(row, cols["hg"]).value = gh
                ws.cell(row, cols["ag"]).value = ga
                if cols["pk"]:
                    ws.cell(row, cols["pk"]).value = pk if pk else None
            wrote += 1
        total += wrote
        if wrote and not dry:
            wb.save(XLSX)
        if not wrote:
            break
    if pens_map and not dry:                       # side file: exact shootout scores for the card
        os.makedirs(os.path.dirname(KO_PENS), exist_ok=True)
        json.dump(pens_map, open(KO_PENS, "w"))
    if unmatched:
        print("  [!] unmatched team names (add to NAME_MAP): " + "; ".join(sorted(unmatched)))
    mode = "added/corrected" if overwrite else "added (blank cells only)"
    if dry:
        print(f"\n  [dry run] {total} result(s) would be {mode}.")
    else:
        print(f"\n  Wrote {total} result(s) into {os.path.basename(XLSX)} ({mode}).")


# --------------------------------------------------------------- offline tests
def _selftest():
    ok = True
    def check(name, got, want):
        nonlocal ok
        flag = "ok " if got == want else "FAIL"
        if got != want: ok = False
        print(f"  [{flag}] {name}: {got!r} (want {want!r})")
    # regular
    check("regular direct", parse_score({"duration": "REGULAR", "winner": "HOME_TEAM",
          "fullTime": {"home": 2, "away": 0}}, True), (2, 0, None))
    check("regular flipped", parse_score({"duration": "REGULAR",
          "fullTime": {"home": 2, "away": 0}}, False), (0, 2, None))
    # extra time decisive
    check("extra time", parse_score({"duration": "EXTRA_TIME", "winner": "AWAY_TEAM",
          "fullTime": {"home": 1, "away": 2}}, True), (1, 2, None))
    # penalty shootout (docs sample: fullTime includes pens 7-6, level 1-1, home won)
    pen = {"duration": "PENALTY_SHOOTOUT", "winner": "HOME_TEAM",
           "fullTime": {"home": 7, "away": 6}, "regularTime": {"home": 1, "away": 1},
           "extraTime": {"home": 0, "away": 0}, "penalties": {"home": 6, "away": 5}}
    check("shootout direct", parse_score(pen, True), (1, 1, "H"))
    check("shootout flipped", parse_score(pen, False), (1, 1, "A"))
    # shootout fallback (no penalties node) uses regular+extra
    check("shootout fallback", parse_score({"duration": "PENALTY_SHOOTOUT", "winner": "AWAY_TEAM",
          "fullTime": {"home": 1, "away": 1}, "regularTime": {"home": 0, "away": 0},
          "extraTime": {"home": 1, "away": 1}}, True), (1, 1, "A"))
    # alt key style (homeTeam/awayTeam)
    check("alt keys", parse_score({"duration": "REGULAR",
          "fullTime": {"homeTeam": 3, "awayTeam": 1}}, True), (3, 1, None))
    # name mapping
    check("name Korea", to_ours("Korea Republic"), "South Korea")
    check("name Turkey", to_ours("Turkey"), "Türkiye")
    check("name USA", to_ours("USA"), "United States")
    check("name accent", to_ours("Cote d'Ivoire"), "Côte d'Ivoire")
    print("\n  SELFTEST", "PASS" if ok else "FAIL")
    return ok


def main():
    if "--selftest" in sys.argv:
        sys.exit(0 if _selftest() else 1)
    try:
        run(dry="--dry-run" in sys.argv, overwrite="--overwrite" in sys.argv)
    except (RuntimeError, PermissionError, OSError) as ex:
        print(f"  [skip] {ex}")


if __name__ == "__main__":
    main()
