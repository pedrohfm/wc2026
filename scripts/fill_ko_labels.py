"""
Fill real team names into the knockout match labels of wc2026_results.xlsx.

The engine maps results by MATCH NUMBER (column A), so the "Home vs Away" text in
column D is only a human aid — for knockout rows it ships as placeholders
("Winner Grp E  vs  3rd-place [slot M74]"). Once the group stage is complete the
Round-of-32 ties are fully determined, so this script resolves them to actual
team names and writes them back, making score entry easy. Re-run it after each
knockout round to fill the next round (R16, QF, …) as feeders are decided.

Third-place slots use the official FIFA allocation (wc2026.config.THIRD_OVERRIDE),
not the engine's approximation, so the eight winner-vs-third ties are correct.

    python scripts/fill_ko_labels.py            # writes labels, saves the file
    python scripts/fill_ko_labels.py --dry-run  # just print what it would write
"""
import os, sys
ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
sys.path.insert(0, os.path.join(ROOT, "src"))
import openpyxl
import wc2026 as E
from wc2026.structure import KO
from wc2026.elo_dynamics import _deterministic_groups

XLSX = os.path.join(ROOT, "wc2026_results.xlsx")
ELO_CSV = os.path.join(ROOT, "wc2026_elo.csv")


def resolve_ko():
    """Return {match_no: (home, away)} for every knockout match whose teams are
       currently determined by the entered results."""
    elo = E.load_elo(ELO_CSV)
    kg, kk = E.load_results(XLSX)
    e = E.apply_known_results(elo, kg, kk, E.THIRD_OVERRIDE)
    res = _deterministic_groups(e, kg)
    if res is None:
        return {}, kg, kk          # group stage not complete -> nothing to resolve
    winners, runners, thirds = res
    thirds.sort(key=lambda x: (x[2]["pts"], x[2]["gd"], x[2]["gf"], e[x[1]]), reverse=True)
    qual = {g: t for g, t, _ in thirds[:8]}
    slot_group = dict(E.THIRD_OVERRIDE) if E.THIRD_OVERRIDE else E.allocate_thirds(list(qual))
    slot_team = {m: qual.get(g) for m, g in slot_group.items()}
    mwin, mlose = {}, {}
    out = {}
    for m in sorted(KO):
        rnd, hs, as_ = KO[m]
        def part(slot):
            typ, ref = slot
            return {"W": winners.get(ref), "RU": runners.get(ref), "3": slot_team.get(ref),
                    "WIN": mwin.get(ref), "LOSE": mlose.get(ref)}[typ]
        home, away = part(hs), part(as_)
        if home and away:
            out[m] = (home, away)
            if m in kk:                       # propagate winners so later rounds resolve too
                ga, gb, pk = kk[m]
                if ga > gb: w = home
                elif gb > ga: w = away
                else: w = home if pk == "H" else (away if pk == "A" else (home if e.get(home, 0) >= e.get(away, 0) else away))
                mwin[m] = w; mlose[m] = away if w == home else home
    return out, kg, kk


def main():
    dry = "--dry-run" in sys.argv
    resolved, kg, kk = resolve_ko()
    if not resolved:
        print(f"  Group stage not complete yet ({len(kg)}/72 entered) — no knockout ties to resolve.")
        return
    wb = openpyxl.load_workbook(XLSX)        # keep formulas/formatting (no data_only)
    ws = wb["Results"]
    # locate the column holding the match number and the matchup label
    hdr = None
    for r in range(1, 12):
        vals = [str(ws.cell(r, c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]
        if any(v.startswith("match") for v in vals):
            hdr = r
            col_match = next(c for c in range(1, ws.max_column + 1)
                             if str(ws.cell(r, c).value or "").strip().lower().startswith("match"))
            # label column = the text column that currently holds "... vs ..."
            col_label = None
            for c in range(1, ws.max_column + 1):
                for rr in range(r + 1, min(r + 80, ws.max_row + 1)):
                    if "vs" in str(ws.cell(rr, c).value or "").lower():
                        col_label = c; break
                if col_label: break
            break
    if hdr is None or col_label is None:
        print("  [!] couldn't locate the match/label columns — file layout unexpected."); return

    n = 0
    for r in range(hdr + 1, ws.max_row + 1):
        mv = ws.cell(r, col_match).value
        try: m = int(float(mv))
        except (TypeError, ValueError): continue
        if m in resolved:
            home, away = resolved[m]
            label = f"{home}  vs  {away}"
            cur = str(ws.cell(r, col_label).value or "")
            if cur != label:
                print(f"  M{m}: {cur!r} -> {label!r}")
                if not dry: ws.cell(r, col_label).value = label
                n += 1
    if dry:
        print(f"\n  [dry run] {n} label(s) would change. Re-run without --dry-run to write.")
        return
    wb.save(XLSX)
    print(f"\n  Wrote {n} knockout label(s) into {os.path.basename(XLSX)}. "
          f"Enter scores in the score cells as matches finish, then run the daily refresh.")


if __name__ == "__main__":
    main()
